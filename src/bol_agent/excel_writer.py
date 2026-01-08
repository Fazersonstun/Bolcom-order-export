"""Excel export functionality with formatting."""

import logging
from pathlib import Path
from datetime import date
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

HEADERS = [
    "export_date",
    "order_id",
    "order_date_time",
    "order_item_id",
    "ean",
    "title",
    "quantity",
    "fulfilment_method",
]


def _ensure_wb(path: Path) -> Tuple[Workbook, Worksheet]:
    """
    Load existing workbook or create new one with formatted headers.

    Args:
        path: Path to Excel file

    Returns:
        Tuple of (Workbook, Worksheet)
    """
    if path.exists():
        logger.debug(f"Loading existing workbook: {path}")
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws

    logger.debug(f"Creating new workbook: {path}")
    wb = Workbook()
    ws = wb.active
    ws.title = "orders"

    # Add headers
    ws.append(HEADERS)

    # Format header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Set column widths
    column_widths = {
        "A": 12,  # export_date
        "B": 15,  # order_id
        "C": 20,  # order_date_time
        "D": 18,  # order_item_id
        "E": 15,  # ean
        "F": 40,  # title
        "G": 10,  # quantity
        "H": 18,  # fulfilment_method
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    logger.debug("Applied header formatting")
    return wb, ws


def append_rows(export_dir: str, rows: List[Dict[str, Any]]) -> Path:
    """
    Append order rows to daily Excel export file.

    Creates a new file if it doesn't exist. Always creates/updates the file
    even if no rows are provided.

    Args:
        export_dir: Directory where Excel files are stored
        rows: List of order item dictionaries to export

    Returns:
        Path to the created/updated Excel file

    Raises:
        IOError: If file cannot be written
    """
    export_path = Path(export_dir) / f"orders_{date.today().isoformat()}.xlsx"
    export_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"Exporting {len(rows)} rows to: {export_path}")

    try:
        wb, ws = _ensure_wb(export_path)

        # Append data rows
        for r in rows:
            ws.append([r.get(h) for h in HEADERS])

        # Save workbook
        wb.save(export_path)
        logger.info(f"Successfully saved Excel file: {export_path}")
        return export_path

    except IOError as e:
        logger.error(f"Failed to save Excel file: {e}")
        raise
